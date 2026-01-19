import os
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

exports_dir = Path("exports")

def formatter(id, response_text):
    data = response_text
    df = pd.DataFrame(data)

    output_dir = exports_dir / id
    os.makedirs(output_dir, exist_ok=True)

    file_name = os.path.join(output_dir, "generated_excel_sheet.xlsx")

    # ------------------ EXPORT WITH PANDAS ------------------

    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Report", index=False)

        ws = writer.sheets["Report"]

        # ------------------ STYLES ------------------

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")  # pro blue
        header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        data_alignment = Alignment(vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # ------------------ FORMAT HEADER ------------------

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[1].height = 30  # padding

        # ------------------ FORMAT DATA ROWS ------------------

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 22
            for cell in row:
                cell.alignment = data_alignment
                cell.border = thin_border

        # ------------------ AUTO COLUMN WIDTH ------------------

        for col_idx, column in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 4

        # ------------------ OPTIONAL PRO FEATURES ------------------

        ws.freeze_panes = "A2"          # sticky header
        ws.auto_filter.ref = ws.dimensions  # filter dropdowns
